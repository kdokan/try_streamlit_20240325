#%%
# デモデータ作成
import pandas as pd

#%%
def clening_google_trend(path_data):
    # データ読み込み
    data = pd.read_csv(
        path_data
        , skiprows=1)

    # カラム名設定
    data = data.rename(
        columns = {
            "週":"ds"
        }
    )

    ## カラム名を正規表現で置換
    data.columns = data.columns.str.replace(r": (日本)", r"_Google検索数")

    # 日付型に変換
    data["ds"] = pd.to_datetime(data["ds"])

    return data



data_buffelin = clening_google_trend("data/buffelin.csv")
data_loxonin = clening_google_trend("data/loxonin.csv")
data_inbound = clening_google_trend("data/inbound.csv")
data_covid = clening_google_trend("data/covid.csv")
data_influenza = clening_google_trend("data/influenza.csv")

data = pd.merge(data_buffelin, data_loxonin, on = "ds", how="left")
data = pd.merge(data, data_inbound, on = "ds", how="left")
data = pd.merge(data, data_covid, on = "ds", how="left")
data = pd.merge(data, data_influenza, on = "ds", how="left")
data = data[data["ds"] >= "2020-01-19"]
data.to_csv('data/input.csv', index=False)  

# %%
y = "バファリン_Google検索数"
data = pd.read_csv('data/input.csv')
data_cor = data.drop(columns="ds")
data_cor = data_cor.corr()[y]
data_cor = data_cor.sort_values(ascending=False)
data_cor = data_cor[data_cor.index != y]
data_cor
# %%
data_cor_top_3 = data_cor.nlargest(2).index
# %%
y = "バファリン_Google検索数"
data_cor_top3_for_plot = data[["ds", y] + list(data_cor_top_3)]
data_cor_top3_for_plot.drop("ds")

# %%
data_cor = data_cor.corr()
#data_cor = data_cor[data_cor[y] != y]
# %%
data = pd.read_excel(
        "data/since2003_visitor_arrivals_February_2024.xlsx"
        , sheet_name = "2024"
        )

data









# %%
from openpyxl import load_workbook

# Excelファイルを読み込む
workbook = load_workbook(filename='ファイル名.xlsx')

# シートを取得
sheet = workbook['シート名']

# 範囲を指定してセルの値を取得
data = []
for row in sheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=3):  # 例: 2行目から10行目まで、1列目から3列目までの範囲
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    data.append(row_data)

# 取得したデータを表示
print(data)